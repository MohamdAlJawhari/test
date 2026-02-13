import csv
import io
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from .config import (
    ALLOWED_CONTACT_EXTENSIONS,
    CONTACTS_METADATA_FILE,
    CONTACTS_PREVIEW_COLUMN_LIMIT,
    CONTACTS_PREVIEW_ROW_LIMIT,
    CONTACTS_UPLOAD_DIR,
)
from .errors import AppError


def is_valid_contacts_filename(filename: str) -> bool:
    suffix = Path(filename).suffix.lower()
    return suffix in ALLOWED_CONTACT_EXTENSIONS


def format_size(num_bytes: int) -> str:
    size = float(max(num_bytes, 0))
    units = ["B", "KB", "MB", "GB"]
    for unit in units:
        if size < 1024 or unit == units[-1]:
            if unit == "B":
                return f"{int(size)} {unit}"
            return f"{size:.1f} {unit}"
        size /= 1024
    return "0 B"


def fallback_display_name(stored_filename: str) -> str:
    filename_parts = stored_filename.split("_", 3)
    return filename_parts[3] if len(filename_parts) == 4 else stored_filename


def sanitize_metadata_text(value, *, max_len: int) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) > max_len:
        return text[:max_len]
    return text

# The following functions handle loading, saving, and managing metadata for uploaded contacts files, as well as validating and parsing the content of those files for use in the application.
def load_contacts_metadata() -> dict[str, dict]:
    if not CONTACTS_METADATA_FILE.exists():
        return {}

    try:
        raw = CONTACTS_METADATA_FILE.read_text(encoding="utf-8")
        parsed = json.loads(raw)
    except (OSError, json.JSONDecodeError):
        return {}

    if not isinstance(parsed, dict):
        return {}

    normalized: dict[str, dict] = {}
    for key, value in parsed.items():
        safe_key = secure_filename(Path(str(key)).name)
        if not safe_key or safe_key != str(key):
            continue

        if not isinstance(value, dict):
            continue

        normalized[safe_key] = {
            "display_name": sanitize_metadata_text(value.get("display_name"), max_len=120),
            "description": sanitize_metadata_text(value.get("description"), max_len=500),
            "created_at": str(value.get("created_at", "")).strip(),
            "updated_at": str(value.get("updated_at", "")).strip(),
        }

    return normalized

# When saving metadata, we ensure the directory exists and handle any file I/O errors by raising an AppError with details for debugging.
def save_contacts_metadata(metadata: dict[str, dict]) -> None:
    CONTACTS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    try:
        CONTACTS_METADATA_FILE.write_text(
            json.dumps(metadata, ensure_ascii=True, indent=2),
            encoding="utf-8",
        )
    except OSError as exc:
        raise AppError(
            code="CONTACTS_METADATA_SAVE_FAILED",
            message="Could not save contacts metadata.",
            status=500,
            details=str(exc),
        ) from exc


def set_contacts_metadata(file_name: str, display_name: str | None = None, description: str | None = None) -> None:
    metadata = load_contacts_metadata()
    now = datetime.now().isoformat(timespec="seconds")
    entry = metadata.get(file_name, {})

    if not entry.get("created_at"):
        entry["created_at"] = now
    entry["updated_at"] = now

    if display_name is not None:
        entry["display_name"] = sanitize_metadata_text(display_name, max_len=120)
    else:
        entry.setdefault("display_name", "")

    if description is not None:
        entry["description"] = sanitize_metadata_text(description, max_len=500)
    else:
        entry.setdefault("description", "")

    metadata[file_name] = entry
    save_contacts_metadata(metadata)


def resolve_saved_contacts_path(file_name: str) -> Path:
    safe_name = secure_filename(Path(str(file_name)).name)
    if not safe_name or safe_name != str(file_name) or not is_valid_contacts_filename(safe_name):
        raise AppError(
            code="INVALID_CONTACT_REFERENCE",
            message="Invalid contacts file reference.",
            status=400,
        )

    file_path = CONTACTS_UPLOAD_DIR / safe_name
    if not file_path.exists() or not file_path.is_file():
        raise AppError(
            code="CONTACT_FILE_NOT_FOUND",
            message="Selected contacts file was not found.",
            status=404,
        )

    return file_path


def remove_contacts_metadata(file_name: str) -> None:
    metadata = load_contacts_metadata()
    if file_name in metadata:
        metadata.pop(file_name, None)
        save_contacts_metadata(metadata)


def build_download_name(file_path: Path, display_name: str) -> str:
    safe_display_name = secure_filename(display_name) or file_path.name
    if Path(safe_display_name).suffix:
        return safe_display_name
    return f"{safe_display_name}{file_path.suffix}"


def save_contacts_upload(filename: str, content: bytes) -> str:
    CONTACTS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = secure_filename(Path(filename).name)
    if not safe_name:
        safe_name = "contacts.xlsx"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    stored_name = f"{timestamp}_{safe_name}"
    destination = CONTACTS_UPLOAD_DIR / stored_name

    try:
        destination.write_bytes(content)
    except OSError as exc:
        raise AppError(
            code="CONTACTS_SAVE_FAILED",
            message="Could not save the uploaded contacts file.",
            status=500,
            details=str(exc),
        ) from exc

    set_contacts_metadata(stored_name, display_name=Path(filename).name, description="")
    return stored_name


def cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = format(value, "f")
        return text.rstrip("0").rstrip(".")
    return str(value).strip()


def decode_csv_bytes(content: bytes) -> str:
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return content.decode("latin-1")


def read_table_from_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = decode_csv_bytes(content)
    raw_rows = list(csv.reader(io.StringIO(text)))
    if not raw_rows:
        return [], []

    raw_headers = [cell_to_text(cell).strip() for cell in raw_rows[0]]
    data_rows = raw_rows[1:]
    max_cols = max(
        len(raw_headers),
        max((len(row) for row in data_rows), default=0),
    )

    if max_cols == 0:
        return [], []

    headers: list[str] = []
    for idx in range(max_cols):
        header_text = raw_headers[idx] if idx < len(raw_headers) else ""
        headers.append(header_text or f"Column {idx + 1}")

    rows: list[list[str]] = []
    for raw_row in data_rows:
        row_values = [
            cell_to_text(raw_row[idx] if idx < len(raw_row) else "")
            for idx in range(max_cols)
        ]
        if any(item.strip() for item in row_values):
            rows.append(row_values)

    return headers, rows


def read_table_from_excel(content: bytes, error_code: str, error_message: str) -> tuple[list[str], list[list[str]]]:
    workbook = None
    try:
        workbook = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise AppError(
            code=error_code,
            message=error_message,
            status=400,
            details=str(exc),
        ) from exc

    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return [], []

        max_cols = max(len(header_row), int(sheet.max_column or 0))
        if max_cols == 0:
            return [], []

        headers: list[str] = []
        for idx in range(max_cols):
            header_text = cell_to_text(header_row[idx] if idx < len(header_row) else "").strip()
            headers.append(header_text or f"Column {idx + 1}")

        rows: list[list[str]] = []
        for raw_row in rows_iter:
            row_values = [
                cell_to_text(raw_row[idx] if idx < len(raw_row) else "")
                for idx in range(max_cols)
            ]
            if any(item.strip() for item in row_values):
                rows.append(row_values)

        return headers, rows
    finally:
        if workbook is not None:
            workbook.close()


def read_contacts_table(
    filename: str,
    content: bytes,
    error_code: str = "CONTACTS_PREVIEW_FAILED",
    error_message: str = "Could not read the selected contacts file.",
) -> tuple[list[str], list[list[str]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        try:
            return read_table_from_csv(content)
        except Exception as exc:
            raise AppError(
                code=error_code,
                message=error_message,
                status=400,
                details=str(exc),
            ) from exc

    return read_table_from_excel(content, error_code=error_code, error_message=error_message)


def validate_contacts_file_for_upload(filename: str, content: bytes) -> None:
    """
    Validate that an uploaded contacts file is readable and has a NUMBERS column.
    This check is used at upload time before storing the file in history.
    """
    headers, _rows = read_contacts_table(
        filename=filename,
        content=content,
        error_code="EXCEL_PARSE_ERROR",
        error_message="Could not read the contacts file. Please check the file format.",
    )

    if not headers:
        raise AppError(
            code="EXCEL_EMPTY",
            message="Contacts file is empty.",
            status=400,
        )

    normalized_headers = [header.upper() for header in headers]
    if "NUMBERS" not in normalized_headers:
        raise AppError(
            code="EXCEL_MISSING_NUMBERS",
            message='Contacts file must contain a "NUMBERS" column in the first row.',
            status=400,
        )


def read_contacts_preview(
    file_path: Path,
    row_limit: int | None = CONTACTS_PREVIEW_ROW_LIMIT,
    column_limit: int | None = CONTACTS_PREVIEW_COLUMN_LIMIT,
) -> dict:
    try:
        content = file_path.read_bytes()
    except OSError as exc:
        raise AppError(
            code="CONTACTS_PREVIEW_FAILED",
            message="Could not read the selected contacts file.",
            status=400,
            details=str(exc),
        ) from exc

    headers, all_rows = read_contacts_table(
        filename=file_path.name,
        content=content,
        error_code="CONTACTS_PREVIEW_FAILED",
        error_message="Could not read the selected contacts file.",
    )

    if not headers:
        return {
            "headers": [],
            "rows": [],
            "total_rows": 0,
            "displayed_rows": 0,
            "truncated": False,
            "total_columns": 0,
            "displayed_columns": 0,
        }

    if column_limit is not None:
        max_cols = max(1, int(column_limit))
        shown_headers = headers[:max_cols]
    else:
        shown_headers = list(headers)

    if row_limit is not None:
        shown_rows_raw = all_rows[: max(0, int(row_limit))]
    else:
        shown_rows_raw = list(all_rows)

    shown_rows = [
        row[: len(shown_headers)] + [""] * max(0, len(shown_headers) - len(row))
        for row in shown_rows_raw
    ]

    truncated = len(all_rows) > len(shown_rows) or len(headers) > len(shown_headers)
    return {
        "headers": shown_headers,
        "rows": shown_rows,
        "total_rows": len(all_rows),
        "displayed_rows": len(shown_rows),
        "truncated": truncated,
        "total_columns": len(headers),
        "displayed_columns": len(shown_headers),
    }


def save_contacts_content(file_path: Path, headers: list, rows: list) -> dict:
    if not isinstance(headers, list):
        raise AppError(
            code="INVALID_CONTACTS_CONTENT",
            message="Headers must be a list.",
            status=400,
        )

    if not isinstance(rows, list):
        raise AppError(
            code="INVALID_CONTACTS_CONTENT",
            message="Rows must be a list.",
            status=400,
        )

    cleaned_headers: list[str] = []
    for idx, header in enumerate(headers):
        header_text = cell_to_text(header).strip()
        cleaned_headers.append(header_text or f"Column {idx + 1}")

    if not cleaned_headers:
        raise AppError(
            code="INVALID_CONTACTS_CONTENT",
            message="Contacts file must have at least one column.",
            status=400,
        )

    normalized_headers = [header.upper() for header in cleaned_headers]
    numbers_columns = [idx for idx, value in enumerate(normalized_headers) if value == "NUMBERS"]
    if len(numbers_columns) != 1:
        raise AppError(
            code="EXCEL_MISSING_NUMBERS",
            message='Contacts file must contain exactly one "NUMBERS" column.',
            status=400,
        )

    numbers_col_idx = numbers_columns[0]
    cleaned_headers[numbers_col_idx] = "NUMBERS"

    duplicates = {
        value for value in normalized_headers if normalized_headers.count(value) > 1
    }
    if duplicates:
        dup_label = ", ".join(sorted(duplicates))
        raise AppError(
            code="INVALID_CONTACTS_CONTENT",
            message=f"Column names must be unique. Duplicate: {dup_label}.",
            status=400,
        )

    cleaned_rows: list[list[str]] = []
    for row in rows:
        if not isinstance(row, list):
            raise AppError(
                code="INVALID_CONTACTS_CONTENT",
                message="Each row must be a list of values.",
                status=400,
            )

        values = [
            cell_to_text(row[idx] if idx < len(row) else "")
            for idx in range(len(cleaned_headers))
        ]
        if any(item.strip() for item in values):
            cleaned_rows.append(values)

    suffix = file_path.suffix.lower()
    try:
        if suffix == ".csv":
            with file_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(cleaned_headers)
                writer.writerows(cleaned_rows)
        else:
            keep_vba = suffix in {".xlsm", ".xltm"}
            workbook = load_workbook(filename=file_path, keep_vba=keep_vba)
            try:
                sheet = workbook.active
                if sheet.max_row > 0:
                    sheet.delete_rows(1, sheet.max_row)
                sheet.append(cleaned_headers)
                for row_values in cleaned_rows:
                    sheet.append(row_values)
                workbook.save(file_path)
            finally:
                workbook.close()
    except AppError:
        raise
    except Exception as exc:
        raise AppError(
            code="CONTACTS_CONTENT_SAVE_FAILED",
            message="Could not save contacts file content.",
            status=500,
            details=str(exc),
        ) from exc

    return {
        "columns": len(cleaned_headers),
        "rows": len(cleaned_rows),
    }


def list_saved_contacts_files() -> list[dict]:
    if not CONTACTS_UPLOAD_DIR.exists():
        return []

    collected: list[tuple[Path, float]] = []
    for path in CONTACTS_UPLOAD_DIR.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() not in ALLOWED_CONTACT_EXTENSIONS:
            continue
        try:
            collected.append((path, path.stat().st_mtime))
        except OSError:
            continue

    metadata = load_contacts_metadata()
    file_names = {path.name for path, _mtime in collected}
    stale_keys = [key for key in metadata.keys() if key not in file_names]
    if stale_keys:
        for key in stale_keys:
            metadata.pop(key, None)
        try:
            save_contacts_metadata(metadata)
        except AppError:
            pass

    files: list[dict] = []
    for path, _mtime in sorted(collected, key=lambda item: item[1], reverse=True):
        try:
            stat = path.stat()
        except OSError:
            continue

        metadata_entry = metadata.get(path.name, {})
        display_name = metadata_entry.get("display_name") or fallback_display_name(path.name)
        description = metadata_entry.get("description", "")

        files.append(
            {
                "name": path.name,
                "display_name": display_name,
                "description": description,
                "modified_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "size_bytes": stat.st_size,
                "size_label": format_size(stat.st_size),
            }
        )

    return files


def parse_contacts_rows(filename: str, content: bytes) -> list[dict]:
    if not is_valid_contacts_filename(filename):
        raise AppError(
            code="INVALID_CONTACTS_FILE",
            message="Upload a valid contacts file (.xlsx or .csv).",
            status=400,
        )

    headers, data_rows = read_contacts_table(
        filename=filename,
        content=content,
        error_code="EXCEL_PARSE_ERROR",
        error_message="Could not read the contacts file. Please check the file format.",
    )

    if not headers:
        raise AppError(
            code="EXCEL_EMPTY",
            message="Contacts file is empty.",
            status=400,
        )

    normalized_headers = [header.upper() for header in headers]
    if "NUMBERS" not in normalized_headers:
        raise AppError(
            code="EXCEL_MISSING_NUMBERS",
            message='Contacts file must contain a "NUMBERS" column in the first row.',
            status=400,
        )

    number_col_idx = normalized_headers.index("NUMBERS")
    rows: list[dict] = []

    for row_idx, row_as_text in enumerate(data_rows, start=2):
        number_value = row_as_text[number_col_idx] if number_col_idx < len(row_as_text) else ""
        if not number_value.strip():
            continue

        row_map: dict[str, str] = {}
        for col_idx, header in enumerate(headers):
            header_name = header.strip()
            if not header_name:
                continue
            value = row_as_text[col_idx] if col_idx < len(row_as_text) else ""
            row_map[header_name] = value
            row_map[header_name.upper()] = value

        row_map["__row_index"] = str(row_idx)
        rows.append(row_map)

    if not rows:
        raise AppError(
            code="EXCEL_NO_ROWS",
            message='No valid rows found. Ensure "NUMBERS" has values.',
            status=400,
        )

    return rows
