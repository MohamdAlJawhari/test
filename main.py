import base64
import csv
import io
import json
import mimetypes
import os
import re
import atexit
import socket
import subprocess
import threading
import time
import webbrowser
from datetime import datetime
from pathlib import Path

import requests
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook
from requests import exceptions as requests_exceptions
from werkzeug.utils import secure_filename

BASE_URL = "http://localhost:3000"
DEFAULT_COUNTRY_CODE = os.getenv("DEFAULT_COUNTRY_CODE", "961")
NODE_PROCESS = None
PLACEHOLDER_PATTERN = re.compile(r"{{\s*([^{}]+?)\s*}}")
BATCH_SEND_DELAY_SECONDS = float(os.getenv("BATCH_SEND_DELAY_SECONDS", "0.8"))
DEFAULT_MESSAGE_TEXT = "Hello {{name}}, your password is {{password}}."
PROJECT_ROOT = Path(__file__).resolve().parent
TEMPLATE_FILE_PATH = PROJECT_ROOT / "template.txt"
CONTACTS_UPLOAD_DIR = PROJECT_ROOT / "data" / "contacts_uploads"
CONTACTS_METADATA_FILE = CONTACTS_UPLOAD_DIR / "metadata.json"
ALLOWED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
ALLOWED_CONTACT_EXTENSIONS = ALLOWED_EXCEL_EXTENSIONS | {".csv"}
CONTACTS_PREVIEW_ROW_LIMIT = int(os.getenv("CONTACTS_PREVIEW_ROW_LIMIT", "20"))
CONTACTS_PREVIEW_COLUMN_LIMIT = int(os.getenv("CONTACTS_PREVIEW_COLUMN_LIMIT", "15"))

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024


class AppError(Exception):
    def __init__(self, code: str, message: str, status: int = 500, details: str | None = None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.status = status
        self.details = details


def _is_valid_contacts_filename(filename: str) -> bool:
    suffix = Path(filename).suffix.lower()
    return suffix in ALLOWED_CONTACT_EXTENSIONS


def _format_size(num_bytes: int) -> str:
    size = float(max(num_bytes, 0))
    units = ["B", "KB", "MB", "GB"]
    for unit in units:
        if size < 1024 or unit == units[-1]:
            if unit == "B":
                return f"{int(size)} {unit}"
            return f"{size:.1f} {unit}"
        size /= 1024
    return "0 B"


def load_default_message_template() -> str:
    try:
        if TEMPLATE_FILE_PATH.exists():
            file_text = TEMPLATE_FILE_PATH.read_text(encoding="utf-8").strip()
            if file_text:
                return file_text
    except OSError:
        pass
    return DEFAULT_MESSAGE_TEXT


def _fallback_display_name(stored_filename: str) -> str:
    filename_parts = stored_filename.split("_", 3)
    return filename_parts[3] if len(filename_parts) == 4 else stored_filename


def _sanitize_metadata_text(value, *, max_len: int) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if len(text) > max_len:
        return text[:max_len]
    return text


def _load_contacts_metadata() -> dict[str, dict]:
    if not CONTACTS_METADATA_FILE.exists():
        return {}

    try:
        raw = CONTACTS_METADATA_FILE.read_text(encoding="utf-8")
        parsed = json.loads(raw)
    except (OSError, json.JSONDecodeError):
        return {}

    if not isinstance(parsed, dict):
        return {}

    normalized: dict[str, dict] = {}
    for key, value in parsed.items():
        safe_key = secure_filename(Path(str(key)).name)
        if not safe_key or safe_key != str(key):
            continue

        if not isinstance(value, dict):
            continue

        normalized[safe_key] = {
            "display_name": _sanitize_metadata_text(value.get("display_name"), max_len=120),
            "description": _sanitize_metadata_text(value.get("description"), max_len=500),
            "created_at": str(value.get("created_at", "")).strip(),
            "updated_at": str(value.get("updated_at", "")).strip(),
        }

    return normalized


def _save_contacts_metadata(metadata: dict[str, dict]) -> None:
    CONTACTS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    try:
        CONTACTS_METADATA_FILE.write_text(
            json.dumps(metadata, ensure_ascii=True, indent=2),
            encoding="utf-8",
        )
    except OSError as exc:
        raise AppError(
            code="CONTACTS_METADATA_SAVE_FAILED",
            message="Could not save contacts metadata.",
            status=500,
            details=str(exc),
        ) from exc


def _set_contacts_metadata(file_name: str, display_name: str | None = None, description: str | None = None) -> None:
    metadata = _load_contacts_metadata()
    now = datetime.now().isoformat(timespec="seconds")
    entry = metadata.get(file_name, {})

    if not entry.get("created_at"):
        entry["created_at"] = now
    entry["updated_at"] = now

    if display_name is not None:
        entry["display_name"] = _sanitize_metadata_text(display_name, max_len=120)
    else:
        entry.setdefault("display_name", "")

    if description is not None:
        entry["description"] = _sanitize_metadata_text(description, max_len=500)
    else:
        entry.setdefault("description", "")

    metadata[file_name] = entry
    _save_contacts_metadata(metadata)


def _resolve_saved_contacts_path(file_name: str) -> Path:
    safe_name = secure_filename(Path(str(file_name)).name)
    if not safe_name or safe_name != str(file_name) or not _is_valid_contacts_filename(safe_name):
        raise AppError(
            code="INVALID_CONTACT_REFERENCE",
            message="Invalid contacts file reference.",
            status=400,
        )

    file_path = CONTACTS_UPLOAD_DIR / safe_name
    if not file_path.exists() or not file_path.is_file():
        raise AppError(
            code="CONTACT_FILE_NOT_FOUND",
            message="Selected contacts file was not found.",
            status=404,
        )

    return file_path


def _remove_contacts_metadata(file_name: str) -> None:
    metadata = _load_contacts_metadata()
    if file_name in metadata:
        metadata.pop(file_name, None)
        _save_contacts_metadata(metadata)


def _build_download_name(file_path: Path, display_name: str) -> str:
    safe_display_name = secure_filename(display_name) or file_path.name
    if Path(safe_display_name).suffix:
        return safe_display_name
    return f"{safe_display_name}{file_path.suffix}"


def save_contacts_upload(filename: str, content: bytes) -> str:
    CONTACTS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    safe_name = secure_filename(Path(filename).name)
    if not safe_name:
        safe_name = "contacts.xlsx"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    stored_name = f"{timestamp}_{safe_name}"
    destination = CONTACTS_UPLOAD_DIR / stored_name

    try:
        destination.write_bytes(content)
    except OSError as exc:
        raise AppError(
            code="CONTACTS_SAVE_FAILED",
            message="Could not save the uploaded contacts file.",
            status=500,
            details=str(exc),
        ) from exc

    _set_contacts_metadata(stored_name, display_name=Path(filename).name, description="")
    return stored_name


def _decode_csv_bytes(content: bytes) -> str:
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return content.decode("latin-1")


def _read_table_from_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = _decode_csv_bytes(content)
    raw_rows = list(csv.reader(io.StringIO(text)))
    if not raw_rows:
        return [], []

    raw_headers = [_cell_to_text(cell).strip() for cell in raw_rows[0]]
    data_rows = raw_rows[1:]
    max_cols = max(
        len(raw_headers),
        max((len(row) for row in data_rows), default=0),
    )

    if max_cols == 0:
        return [], []

    headers: list[str] = []
    for idx in range(max_cols):
        header_text = raw_headers[idx] if idx < len(raw_headers) else ""
        headers.append(header_text or f"Column {idx + 1}")

    rows: list[list[str]] = []
    for raw_row in data_rows:
        row_values = [
            _cell_to_text(raw_row[idx] if idx < len(raw_row) else "")
            for idx in range(max_cols)
        ]
        if any(item.strip() for item in row_values):
            rows.append(row_values)

    return headers, rows


def _read_table_from_excel(content: bytes, error_code: str, error_message: str) -> tuple[list[str], list[list[str]]]:
    workbook = None
    try:
        workbook = load_workbook(filename=io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise AppError(
            code=error_code,
            message=error_message,
            status=400,
            details=str(exc),
        ) from exc

    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return [], []

        max_cols = max(len(header_row), int(sheet.max_column or 0))
        if max_cols == 0:
            return [], []

        headers: list[str] = []
        for idx in range(max_cols):
            header_text = _cell_to_text(header_row[idx] if idx < len(header_row) else "").strip()
            headers.append(header_text or f"Column {idx + 1}")

        rows: list[list[str]] = []
        for raw_row in rows_iter:
            row_values = [
                _cell_to_text(raw_row[idx] if idx < len(raw_row) else "")
                for idx in range(max_cols)
            ]
            if any(item.strip() for item in row_values):
                rows.append(row_values)

        return headers, rows
    finally:
        if workbook is not None:
            workbook.close()


def _read_contacts_table(
    filename: str,
    content: bytes,
    error_code: str = "CONTACTS_PREVIEW_FAILED",
    error_message: str = "Could not read the selected contacts file.",
) -> tuple[list[str], list[list[str]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        try:
            return _read_table_from_csv(content)
        except Exception as exc:
            raise AppError(
                code=error_code,
                message=error_message,
                status=400,
                details=str(exc),
            ) from exc

    return _read_table_from_excel(content, error_code=error_code, error_message=error_message)


def read_contacts_preview(
    file_path: Path,
    row_limit: int | None = CONTACTS_PREVIEW_ROW_LIMIT,
    column_limit: int | None = CONTACTS_PREVIEW_COLUMN_LIMIT,
) -> dict:
    try:
        content = file_path.read_bytes()
    except OSError as exc:
        raise AppError(
            code="CONTACTS_PREVIEW_FAILED",
            message="Could not read the selected contacts file.",
            status=400,
            details=str(exc),
        ) from exc

    headers, all_rows = _read_contacts_table(
        filename=file_path.name,
        content=content,
        error_code="CONTACTS_PREVIEW_FAILED",
        error_message="Could not read the selected contacts file.",
    )

    if not headers:
        return {
            "headers": [],
            "rows": [],
            "total_rows": 0,
            "displayed_rows": 0,
            "truncated": False,
            "total_columns": 0,
            "displayed_columns": 0,
        }

    if column_limit is not None:
        max_cols = max(1, int(column_limit))
        shown_headers = headers[:max_cols]
    else:
        shown_headers = list(headers)

    if row_limit is not None:
        shown_rows_raw = all_rows[: max(0, int(row_limit))]
    else:
        shown_rows_raw = list(all_rows)

    shown_rows = [
        row[: len(shown_headers)] + [""] * max(0, len(shown_headers) - len(row))
        for row in shown_rows_raw
    ]

    truncated = len(all_rows) > len(shown_rows) or len(headers) > len(shown_headers)
    return {
        "headers": shown_headers,
        "rows": shown_rows,
        "total_rows": len(all_rows),
        "displayed_rows": len(shown_rows),
        "truncated": truncated,
        "total_columns": len(headers),
        "displayed_columns": len(shown_headers),
    }


def save_contacts_content(file_path: Path, headers: list, rows: list) -> dict:
    if not isinstance(headers, list):
        raise AppError(
            code="INVALID_CONTACTS_CONTENT",
            message="Headers must be a list.",
            status=400,
        )

    if not isinstance(rows, list):
        raise AppError(
            code="INVALID_CONTACTS_CONTENT",
            message="Rows must be a list.",
            status=400,
        )

    cleaned_headers: list[str] = []
    for idx, header in enumerate(headers):
        header_text = _cell_to_text(header).strip()
        cleaned_headers.append(header_text or f"Column {idx + 1}")

    if not cleaned_headers:
        raise AppError(
            code="INVALID_CONTACTS_CONTENT",
            message="Contacts file must have at least one column.",
            status=400,
        )

    normalized_headers = [header.upper() for header in cleaned_headers]
    numbers_columns = [idx for idx, value in enumerate(normalized_headers) if value == "NUMBERS"]
    if len(numbers_columns) != 1:
        raise AppError(
            code="EXCEL_MISSING_NUMBERS",
            message='Contacts file must contain exactly one "NUMBERS" column.',
            status=400,
        )

    numbers_col_idx = numbers_columns[0]
    cleaned_headers[numbers_col_idx] = "NUMBERS"

    duplicates = {
        value for value in normalized_headers if normalized_headers.count(value) > 1
    }
    if duplicates:
        dup_label = ", ".join(sorted(duplicates))
        raise AppError(
            code="INVALID_CONTACTS_CONTENT",
            message=f"Column names must be unique. Duplicate: {dup_label}.",
            status=400,
        )

    cleaned_rows: list[list[str]] = []
    for row in rows:
        if not isinstance(row, list):
            raise AppError(
                code="INVALID_CONTACTS_CONTENT",
                message="Each row must be a list of values.",
                status=400,
            )

        values = [
            _cell_to_text(row[idx] if idx < len(row) else "")
            for idx in range(len(cleaned_headers))
        ]
        if any(item.strip() for item in values):
            cleaned_rows.append(values)

    suffix = file_path.suffix.lower()
    try:
        if suffix == ".csv":
            with file_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(cleaned_headers)
                writer.writerows(cleaned_rows)
        else:
            keep_vba = suffix in {".xlsm", ".xltm"}
            workbook = load_workbook(filename=file_path, keep_vba=keep_vba)
            try:
                sheet = workbook.active
                if sheet.max_row > 0:
                    sheet.delete_rows(1, sheet.max_row)
                sheet.append(cleaned_headers)
                for row_values in cleaned_rows:
                    sheet.append(row_values)
                workbook.save(file_path)
            finally:
                workbook.close()
    except AppError:
        raise
    except Exception as exc:
        raise AppError(
            code="CONTACTS_CONTENT_SAVE_FAILED",
            message="Could not save contacts file content.",
            status=500,
            details=str(exc),
        ) from exc

    return {
        "columns": len(cleaned_headers),
        "rows": len(cleaned_rows),
    }


def list_saved_contacts_files() -> list[dict]:
    if not CONTACTS_UPLOAD_DIR.exists():
        return []

    collected: list[tuple[Path, float]] = []
    for path in CONTACTS_UPLOAD_DIR.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() not in ALLOWED_CONTACT_EXTENSIONS:
            continue
        try:
            collected.append((path, path.stat().st_mtime))
        except OSError:
            continue

    metadata = _load_contacts_metadata()
    file_names = {path.name for path, _mtime in collected}
    stale_keys = [key for key in metadata.keys() if key not in file_names]
    if stale_keys:
        for key in stale_keys:
            metadata.pop(key, None)
        try:
            _save_contacts_metadata(metadata)
        except AppError:
            pass

    files: list[dict] = []
    for path, _mtime in sorted(collected, key=lambda item: item[1], reverse=True):
        try:
            stat = path.stat()
        except OSError:
            continue

        metadata_entry = metadata.get(path.name, {})
        display_name = metadata_entry.get("display_name") or _fallback_display_name(path.name)
        description = metadata_entry.get("description", "")

        files.append(
            {
                "name": path.name,
                "display_name": display_name,
                "description": description,
                "modified_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "size_bytes": stat.st_size,
                "size_label": _format_size(stat.st_size),
            }
        )

    return files


def _json_error(code: str, message: str, status: int, details: str | None = None):
    payload = {"ok": False, "error_code": code, "error": message}
    if details:
        payload["details"] = details
    return jsonify(payload), status


def _handle_api_exception(exc: Exception):
    if isinstance(exc, AppError):
        return _json_error(exc.code, exc.message, exc.status, exc.details)
    return _json_error(
        code="UNEXPECTED_SERVER_ERROR",
        message="Unexpected server error. Please try again.",
        status=500,
        details=str(exc),
    )


def _map_node_error(status_code: int, raw_error: str | None = None) -> AppError:
    details = raw_error or ""
    lowered = details.lower()

    if "not ready" in lowered or status_code == 503:
        return AppError(
            code="WHATSAPP_NOT_READY",
            message="WhatsApp is not ready yet. Keep the QR window open and scan again.",
            status=503,
            details=details,
        )

    if "timed out" in lowered:
        return AppError(
            code="DELIVERY_TIMEOUT",
            message="The connection is slow. Sending is taking longer than expected. It may still arrive.",
            status=504,
            details=details,
        )

    if "ack failed" in lowered:
        return AppError(
            code="DELIVERY_FAILED",
            message="Message delivery failed. Please verify the number and try again.",
            status=502,
            details=details,
        )

    if status_code == 413:
        return AppError(
            code="MEDIA_TOO_LARGE",
            message="Selected media is too large for this request. Try a smaller file.",
            status=413,
            details=details,
        )

    return AppError(
        code="WHATSAPP_API_ERROR",
        message="WhatsApp service returned an error. Please try again.",
        status=max(status_code, 500),
        details=details,
    )


def normalize_to_whatsapp_id(raw_phone: str) -> str:
    """
    Convert user input into WhatsApp ID format expected by WPPConnect.
    Output example: 96181744432@c.us
    """
    if not raw_phone or not raw_phone.strip():
        raise ValueError("Phone number is required.")

    value = raw_phone.strip()
    if value.endswith("@c.us"):
        value = value[:-5]

    value = (
        value.replace(" ", "")
        .replace("-", "")
        .replace("(", "")
        .replace(")", "")
    )

    explicit_international = value.startswith("+") or value.startswith("00")
    if value.startswith("+"):
        value = value[1:]
    elif value.startswith("00"):
        value = value[2:]

    digits = "".join(ch for ch in value if ch.isdigit())
    if not digits:
        raise ValueError("Phone number must contain digits.")

    country_digits = "".join(ch for ch in DEFAULT_COUNTRY_CODE if ch.isdigit())
    if not country_digits:
        raise ValueError("DEFAULT_COUNTRY_CODE must contain digits.")

    if explicit_international or digits.startswith(country_digits):
        normalized_digits = digits
    else:
        local_digits = digits.lstrip("0")
        if not local_digits:
            raise ValueError("Phone number is invalid.")
        normalized_digits = f"{country_digits}{local_digits}"

    return f"{normalized_digits}@c.us"


def upload_to_data_url(uploaded_file) -> str:
    raw = uploaded_file.read()
    if not raw:
        raise ValueError("Selected media file is empty.")

    mime_type = uploaded_file.mimetype
    if not mime_type:
        mime_type, _ = mimetypes.guess_type(uploaded_file.filename or "")
    if not mime_type:
        mime_type = "application/octet-stream"

    b64 = base64.b64encode(raw).decode("utf-8")
    return f"data:{mime_type};base64,{b64}"


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = format(value, "f")
        return text.rstrip("0").rstrip(".")
    return str(value).strip()


def parse_excel_contacts(filename: str, content: bytes) -> list[dict]:
    if not _is_valid_contacts_filename(filename):
        raise AppError(
            code="INVALID_CONTACTS_FILE",
            message="Upload a valid contacts file (.xlsx or .csv).",
            status=400,
        )

    headers, data_rows = _read_contacts_table(
        filename=filename,
        content=content,
        error_code="EXCEL_PARSE_ERROR",
        error_message="Could not read the contacts file. Please check the file format.",
    )

    if not headers:
        raise AppError(
            code="EXCEL_EMPTY",
            message="Contacts file is empty.",
            status=400,
        )

    normalized_headers = [header.upper() for header in headers]
    if "NUMBERS" not in normalized_headers:
        raise AppError(
            code="EXCEL_MISSING_NUMBERS",
            message='Contacts file must contain a "NUMBERS" column in the first row.',
            status=400,
        )

    number_col_idx = normalized_headers.index("NUMBERS")
    rows: list[dict] = []

    for row_idx, row_as_text in enumerate(data_rows, start=2):
        number_value = row_as_text[number_col_idx] if number_col_idx < len(row_as_text) else ""
        if not number_value.strip():
            continue

        row_map: dict[str, str] = {}
        for col_idx, header in enumerate(headers):
            header_name = header.strip()
            if not header_name:
                continue
            value = row_as_text[col_idx] if col_idx < len(row_as_text) else ""
            row_map[header_name] = value
            row_map[header_name.upper()] = value

        row_map["__row_index"] = str(row_idx)
        rows.append(row_map)

    if not rows:
        raise AppError(
            code="EXCEL_NO_ROWS",
            message='No valid rows found. Ensure "NUMBERS" has values.',
            status=400,
        )

    return rows


def render_message_template(template: str, row_map: dict) -> str:
    if not template:
        return ""

    lookup = {str(key).strip().lower(): str(value) for key, value in row_map.items()}
    missing_keys: set[str] = set()

    def replace(match: re.Match) -> str:
        raw_key = match.group(1).strip()
        key = raw_key.lower()
        if key in lookup:
            return lookup[key]
        missing_keys.add(raw_key)
        return match.group(0)

    rendered = PLACEHOLDER_PATTERN.sub(replace, template)
    if missing_keys:
        row_ref = row_map.get("__row_index", "?")
        raise AppError(
            code="MISSING_TEMPLATE_VARIABLE",
            message=f"Missing column(s) in contacts file for placeholders: {', '.join(sorted(missing_keys))}.",
            status=400,
            details=f"Row {row_ref}",
        )

    return rendered


def call_node_api(method: str, endpoint: str, payload: dict | None = None, timeout: int = 10) -> dict:
    method = method.upper().strip()
    url = f"{BASE_URL}{endpoint}"

    try:
        if method == "GET":
            response = requests.get(url, timeout=timeout)
        elif method == "POST":
            response = requests.post(url, json=payload, timeout=timeout)
        else:
            raise AppError(
                code="INVALID_METHOD",
                message="Internal server configuration error.",
                status=500,
                details=f"Unsupported HTTP method: {method}",
            )
    except requests_exceptions.Timeout as exc:
        raise AppError(
            code="NODE_API_TIMEOUT",
            message="WhatsApp service is taking too long to respond. Check your connection and try again.",
            status=504,
            details=str(exc),
        ) from exc
    except requests_exceptions.ConnectionError as exc:
        raise AppError(
            code="NODE_API_UNREACHABLE",
            message="Cannot connect to WhatsApp service. Wait a few seconds and try again.",
            status=503,
            details=str(exc),
        ) from exc
    except requests_exceptions.RequestException as exc:
        raise AppError(
            code="NODE_API_REQUEST_ERROR",
            message="Network error while contacting WhatsApp service.",
            status=502,
            details=str(exc),
        ) from exc

    try:
        data = response.json()
    except ValueError as exc:
        raise AppError(
            code="NODE_API_BAD_RESPONSE",
            message="WhatsApp service returned an unexpected response.",
            status=502,
            details=str(exc),
        ) from exc

    if not response.ok:
        raw_error = data.get("error") if isinstance(data, dict) else None
        raise _map_node_error(status_code=response.status_code, raw_error=raw_error)

    return data


def post_json(endpoint: str, payload: dict, timeout: int) -> dict:
    data = call_node_api("POST", endpoint, payload=payload, timeout=timeout)
    if not data.get("ok"):
        raise _map_node_error(status_code=502, raw_error=data.get("error"))
    return data


def send_text(to: str, message: str, keep_session: bool = False) -> None:
    post_json(
        endpoint="/send-text",
        payload={"to": to, "message": message, "keepSession": keep_session},
        timeout=10,
    )


def send_media(
    to: str,
    filename: str,
    caption: str,
    data_url: str,
    keep_session: bool = False,
) -> None:
    post_json(
        endpoint="/send-media",
        payload={
            "to": to,
            "filename": filename,
            "caption": caption,
            "base64": data_url,
            "keepSession": keep_session,
        },
        timeout=120,
    )


def logout_session() -> None:
    data = call_node_api("POST", "/session/logout", payload={}, timeout=20)
    if not data.get("ok"):
        raise AppError(
            code="LOGOUT_FAILED",
            message="Failed to close WhatsApp session after sending.",
            status=502,
            details=data.get("error"),
        )


@app.get("/")
def home():
    uploaded_excel_files = list_saved_contacts_files()
    selected_existing_contacts_file = ""
    requested_file = request.args.get("existing_contacts_file", "").strip()
    if requested_file:
        try:
            selected_existing_contacts_file = _resolve_saved_contacts_path(requested_file).name
        except AppError:
            selected_existing_contacts_file = ""

    return render_template(
        "index.html",
        default_country_code=DEFAULT_COUNTRY_CODE,
        default_message_template=load_default_message_template(),
        uploaded_excel_files=uploaded_excel_files,
        selected_existing_contacts_file=selected_existing_contacts_file,
    )


@app.post("/api/auth/start")
def api_auth_start():
    try:
        data = call_node_api("POST", "/auth/start", payload={}, timeout=10)
        if not data.get("ok"):
            raise AppError(
                code="AUTH_START_FAILED",
                message="Failed to start WhatsApp login.",
                status=502,
                details=data.get("error"),
            )
        return jsonify(data)
    except Exception as exc:
        return _handle_api_exception(exc)


@app.get("/api/auth/status")
def api_auth_status():
    try:
        data = call_node_api("GET", "/auth/status", timeout=10)
        if not data.get("ok"):
            raise AppError(
                code="AUTH_STATUS_FAILED",
                message="Failed to read WhatsApp login status.",
                status=502,
                details=data.get("error"),
            )
        return jsonify(data)
    except Exception as exc:
        return _handle_api_exception(exc)


@app.get("/api/contacts/history")
def api_contacts_history():
    try:
        return jsonify({"ok": True, "files": list_saved_contacts_files()})
    except Exception as exc:
        return _handle_api_exception(exc)


@app.post("/api/contacts/upload")
def api_contacts_upload():
    contacts_file = request.files.get("contacts_file")

    try:
        if not contacts_file or not (contacts_file.filename or "").strip():
            raise AppError(
                code="CONTACTS_UPLOAD_MISSING",
                message="Choose a contacts file to upload.",
                status=400,
            )

        contacts_filename = (contacts_file.filename or "").strip()
        if not _is_valid_contacts_filename(contacts_filename):
            raise AppError(
                code="INVALID_CONTACTS_FILE",
                message="Upload a valid contacts file (.xlsx or .csv).",
                status=400,
            )

        contacts_bytes = contacts_file.read()
        if not contacts_bytes:
            raise AppError(
                code="EXCEL_EMPTY",
                message="Contacts file is empty.",
                status=400,
            )

        # Validate that the file can be parsed before storing it.
        _read_contacts_table(
            filename=contacts_filename,
            content=contacts_bytes,
            error_code="EXCEL_PARSE_ERROR",
            error_message="Could not read the contacts file. Please check the file format.",
        )

        stored_name = save_contacts_upload(contacts_filename, contacts_bytes)
        file_info = next(
            (item for item in list_saved_contacts_files() if item.get("name") == stored_name),
            None,
        )

        return jsonify(
            {
                "ok": True,
                "message": "Contacts file uploaded and saved.",
                "file": file_info
                or {
                    "name": stored_name,
                    "display_name": Path(contacts_filename).name,
                    "description": "",
                    "modified_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "size_bytes": len(contacts_bytes),
                    "size_label": _format_size(len(contacts_bytes)),
                },
            }
        )
    except Exception as exc:
        return _handle_api_exception(exc)


@app.get("/api/contacts/<path:file_name>/preview")
def api_contacts_preview(file_name: str):
    try:
        file_path = _resolve_saved_contacts_path(file_name)
        metadata = _load_contacts_metadata().get(file_path.name, {})
        preview = read_contacts_preview(file_path)
        return jsonify(
            {
                "ok": True,
                "file": {
                    "name": file_path.name,
                    "display_name": metadata.get("display_name") or _fallback_display_name(file_path.name),
                    "description": metadata.get("description", ""),
                },
                "preview": preview,
            }
        )
    except Exception as exc:
        return _handle_api_exception(exc)


@app.post("/api/contacts/<path:file_name>/content")
def api_contacts_content(file_name: str):
    try:
        file_path = _resolve_saved_contacts_path(file_name)
        payload = request.get_json(silent=True) or {}
        if not isinstance(payload, dict):
            payload = {}

        headers = payload.get("headers", [])
        rows = payload.get("rows", [])
        summary = save_contacts_content(file_path=file_path, headers=headers, rows=rows)
        preview = read_contacts_preview(file_path, row_limit=None, column_limit=None)
        return jsonify({"ok": True, "summary": summary, "preview": preview})
    except Exception as exc:
        return _handle_api_exception(exc)


@app.post("/api/contacts/<path:file_name>/metadata")
def api_contacts_metadata(file_name: str):
    try:
        _ = _resolve_saved_contacts_path(file_name)
        payload = request.get_json(silent=True) or {}
        if not isinstance(payload, dict):
            payload = {}

        display_name = _sanitize_metadata_text(payload.get("display_name"), max_len=120)
        description = _sanitize_metadata_text(payload.get("description"), max_len=500)
        if not display_name:
            raise AppError(
                code="INVALID_CONTACT_METADATA",
                message="Contacts file name cannot be empty.",
                status=400,
            )

        _set_contacts_metadata(file_name, display_name=display_name, description=description)
        return jsonify({"ok": True})
    except Exception as exc:
        return _handle_api_exception(exc)


@app.get("/api/contacts/<path:file_name>/download")
def api_contacts_download(file_name: str):
    try:
        file_path = _resolve_saved_contacts_path(file_name)
        metadata = _load_contacts_metadata().get(file_path.name, {})
        display_name = metadata.get("display_name") or _fallback_display_name(file_path.name)
        download_name = _build_download_name(file_path, display_name)
        return send_file(
            file_path,
            as_attachment=True,
            download_name=download_name,
            mimetype=mimetypes.guess_type(str(file_path))[0] or "application/octet-stream",
        )
    except Exception as exc:
        return _handle_api_exception(exc)


@app.delete("/api/contacts/<path:file_name>")
def api_contacts_delete(file_name: str):
    try:
        file_path = _resolve_saved_contacts_path(file_name)
        try:
            file_path.unlink()
        except OSError as exc:
            raise AppError(
                code="CONTACTS_DELETE_FAILED",
                message="Could not delete contacts file.",
                status=500,
                details=str(exc),
            ) from exc

        _remove_contacts_metadata(file_path.name)
        return jsonify({"ok": True, "message": "Contacts file deleted."})
    except Exception as exc:
        return _handle_api_exception(exc)


@app.get("/contacts/<path:file_name>")
def contacts_details_page(file_name: str):
    try:
        file_path = _resolve_saved_contacts_path(file_name)
        file_info = next(
            (item for item in list_saved_contacts_files() if item.get("name") == file_path.name),
            None,
        )
        if not file_info:
            raise AppError(
                code="CONTACT_FILE_NOT_FOUND",
                message="Selected contacts file was not found.",
                status=404,
            )

        preview = read_contacts_preview(file_path, row_limit=None, column_limit=None)
        return render_template(
            "contact_details.html",
            contact_file=file_info,
            preview=preview,
            load_error="",
        )
    except AppError as exc:
        return (
            render_template(
                "contact_details.html",
                contact_file=None,
                preview={
                    "headers": [],
                    "rows": [],
                    "total_rows": 0,
                    "displayed_rows": 0,
                    "truncated": False,
                    "total_columns": 0,
                    "displayed_columns": 0,
                },
                load_error=exc.message,
            ),
            exc.status,
        )
    except Exception:
        return (
            render_template(
                "contact_details.html",
                contact_file=None,
                preview={
                    "headers": [],
                    "rows": [],
                    "total_rows": 0,
                    "displayed_rows": 0,
                    "truncated": False,
                    "total_columns": 0,
                    "displayed_columns": 0,
                },
                load_error="Could not load this contacts file.",
            ),
            500,
        )


@app.post("/api/send")
def api_send():
    phone = request.form.get("phone", "").strip()
    message = request.form.get("message", "").strip()
    uploaded_file = request.files.get("media")
    contacts_file = request.files.get("contacts_file")
    existing_contacts_file = request.form.get("existing_contacts_file", "").strip()

    try:
        has_contacts_file = bool(contacts_file and contacts_file.filename)
        has_existing_contacts_file = bool(existing_contacts_file)
        has_media = bool(uploaded_file and uploaded_file.filename)
        has_message = bool(message)

        if not has_contacts_file and not has_existing_contacts_file and not phone:
            raise AppError(
                code="MISSING_TARGET",
                message="Enter a phone number, upload a contacts file, or select an existing contacts file.",
                status=400,
            )

        if not has_message and not has_media:
            raise AppError(
                code="MISSING_CONTENT",
                message="Write a message or choose a media file.",
                status=400,
            )

        if has_contacts_file or has_existing_contacts_file:
            contacts_filename = ""
            contacts_bytes = b""

            if has_contacts_file:
                contacts_filename = (contacts_file.filename or "").strip()
                if not contacts_filename:
                    raise AppError(
                        code="INVALID_CONTACTS_FILE",
                        message="Upload a valid contacts file (.xlsx or .csv).",
                        status=400,
                    )

                if not _is_valid_contacts_filename(contacts_filename):
                    raise AppError(
                        code="INVALID_CONTACTS_FILE",
                        message="Upload a valid contacts file (.xlsx or .csv).",
                        status=400,
                    )

                contacts_bytes = contacts_file.read()
                if not contacts_bytes:
                    raise AppError(
                        code="EXCEL_EMPTY",
                        message="Contacts file is empty.",
                        status=400,
                    )

                save_contacts_upload(contacts_filename, contacts_bytes)
            else:
                existing_path = _resolve_saved_contacts_path(existing_contacts_file)
                contacts_filename = existing_path.name
                try:
                    contacts_bytes = existing_path.read_bytes()
                except OSError as exc:
                    raise AppError(
                        code="CONTACTS_PREVIEW_FAILED",
                        message="Could not read the selected contacts file.",
                        status=400,
                        details=str(exc),
                    ) from exc

                if not contacts_bytes:
                    raise AppError(
                        code="EXCEL_EMPTY",
                        message="Contacts file is empty.",
                        status=400,
                    )

            rows = parse_excel_contacts(contacts_filename, contacts_bytes)
            media_data_url = upload_to_data_url(uploaded_file) if has_media else ""
            media_filename = uploaded_file.filename if has_media else ""

            try:
                for index, row in enumerate(rows):
                    row_number = row.get("NUMBERS", "")
                    to = normalize_to_whatsapp_id(row_number)
                    rendered_message = render_message_template(message, row).strip()
                    keep_session = True

                    if has_media:
                        send_media(
                            to=to,
                            filename=media_filename,
                            caption=rendered_message,
                            data_url=media_data_url,
                            keep_session=keep_session,
                        )
                    else:
                        if not rendered_message:
                            raise AppError(
                                code="EMPTY_ROW_MESSAGE",
                                message="Message is empty for one or more rows after variable replacement.",
                                status=400,
                                details=f"Row {row.get('__row_index', '?')}",
                            )
                        send_text(to=to, message=rendered_message, keep_session=keep_session)

                    if index < len(rows) - 1 and BATCH_SEND_DELAY_SECONDS > 0:
                        time.sleep(BATCH_SEND_DELAY_SECONDS)
            except AppError:
                try:
                    logout_session()
                except Exception:
                    pass
                raise
            except Exception as exc:
                try:
                    logout_session()
                except Exception:
                    pass
                raise AppError(
                    code="BATCH_ROW_FAILED",
                    message="Failed while sending one of the contacts rows.",
                    status=502,
                    details=str(exc),
                ) from exc

            logout_warning = None
            try:
                logout_session()
            except Exception as exc:
                logout_warning = str(exc)

            result_message = (
                f"Batch sent successfully to {len(rows)} row(s). "
                "You are logged out; a new QR code will be required next send."
            )
            if logout_warning:
                result_message = (
                    f"Batch sent successfully to {len(rows)} row(s), but automatic logout failed. "
                    "Restart the app before the next batch."
                )

            return jsonify(
                {
                    "ok": True,
                    "message": result_message,
                }
            )

        to = normalize_to_whatsapp_id(phone)

        if has_media:
            data_url = upload_to_data_url(uploaded_file)
            send_media(
                to=to,
                filename=uploaded_file.filename,
                caption=message,
                data_url=data_url,
                keep_session=False,
            )
            return jsonify(
                {
                    "ok": True,
                    "message": f"Media sent to {to}. You are logged out; a new QR code will be required next send.",
                }
            )

        send_text(to=to, message=message, keep_session=False)
        return jsonify(
            {
                "ok": True,
                "message": f"Text sent to {to}. You are logged out; a new QR code will be required next send.",
            }
        )
    except ValueError as exc:
        return _json_error(
            code="VALIDATION_ERROR",
            message=str(exc),
            status=400,
        )
    except Exception as exc:
        return _handle_api_exception(exc)


@app.errorhandler(413)
def handle_payload_too_large(_exc):
    return _json_error(
        code="MEDIA_TOO_LARGE",
        message="Selected media is too large. Choose a smaller file and try again.",
        status=413,
    )


def is_port_available(port: int, host: str = "127.0.0.1") -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        try:
            sock.bind((host, port))
            return True
        except OSError:
            return False


def find_available_port(preferred_port: int) -> int:
    if preferred_port > 0 and is_port_available(preferred_port):
        return preferred_port

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def start_node_server(api_port: int) -> subprocess.Popen:
    env = os.environ.copy()
    env["API_PORT"] = str(api_port)

    project_root = Path(__file__).resolve().parent
    try:
        process = subprocess.Popen(
            ["node", "index.js"],
            cwd=project_root,
            env=env,
        )
    except FileNotFoundError as exc:
        raise RuntimeError(
            'Node.js is not available in PATH. Install Node.js or run "node index.js" manually.'
        ) from exc

    return process


def stop_node_server() -> None:
    global NODE_PROCESS
    process = NODE_PROCESS
    NODE_PROCESS = None

    if not process:
        return

    if process.poll() is not None:
        return

    process.terminate()
    try:
        process.wait(timeout=5)
    except subprocess.TimeoutExpired:
        process.kill()


def wait_for_node_ready(base_url: str, timeout_seconds: int = 20) -> None:
    deadline = time.time() + timeout_seconds
    last_error = None

    while time.time() < deadline:
        if NODE_PROCESS and NODE_PROCESS.poll() is not None:
            raise RuntimeError("Node API process exited before becoming ready.")

        try:
            response = requests.get(f"{base_url}/auth/status", timeout=1.5)
            if response.ok:
                return
        except requests.RequestException as exc:
            last_error = exc

        time.sleep(0.4)

    raise RuntimeError(
        f"Node API did not become ready on {base_url} within {timeout_seconds} seconds."
        + (f" Last error: {last_error}" if last_error else "")
    )


def open_browser_soon(url: str, delay_seconds: float = 0.8) -> None:
    timer = threading.Timer(delay_seconds, lambda: webbrowser.open(url))
    timer.daemon = True
    timer.start()


if __name__ == "__main__":
    preferred_api_port = int(os.getenv("API_PORT", "3000"))
    api_port = find_available_port(preferred_api_port)

    BASE_URL = f"http://127.0.0.1:{api_port}"
    NODE_PROCESS = start_node_server(api_port)
    atexit.register(stop_node_server)
    wait_for_node_ready(BASE_URL)

    preferred_ui_port = int(os.getenv("FLASK_PORT", "5000"))
    ui_port = find_available_port(preferred_ui_port)
    debug_mode = os.getenv("FLASK_DEBUG", "1") == "1"
    ui_url = f"http://127.0.0.1:{ui_port}"

    print(f"Node API running on: {BASE_URL}")
    print(f"Opening UI on: {ui_url}")
    open_browser_soon(ui_url)

    try:
        app.run(
            host="127.0.0.1",
            port=ui_port,
            debug=debug_mode,
            use_reloader=False,
        )
    finally:
        stop_node_server()
